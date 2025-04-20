"""
实验结果Excel记录模块 - 将实验结果记录到Excel文件中
"""

import os
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

def log_experiment_to_excel(args, train_best_accuracy, test_accuracy, label_to_idx, 
                          tcn_history=None, metrics=None, excel_file="experiment_results.xlsx"):
    """
    将实验结果记录到Excel文件中，每次实验添加一行
    
    参数:
    args - 命令行参数
    train_best_accuracy - 训练中获得的最佳准确率
    test_accuracy - 测试集上的准确率
    label_to_idx - 标签到索引的映射（用于记录类别数量）
    tcn_history - TCN训练历史数据 (可选)
    metrics - 更多的评估指标 (可选)
    excel_file - Excel文件路径
    """
    # 创建当前实验的数据字典
    experiment_data = {
        '实验时间': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        '训练最佳准确率(%)': round(train_best_accuracy, 2),
        '测试准确率(%)': round(test_accuracy, 2),
        '准确率差异(%)': round(train_best_accuracy - test_accuracy, 2),
        '训练轮数': args.epochs,
        '学习率': args.lr,
        '类别数量': len(label_to_idx),
        '分类器类型': args.classifier if hasattr(args, 'classifier') else 'svm',
        '时域特征权重': args.time_weight,
        '频域特征权重': args.freq_weight,
        '使用时域特征': args.use_time,
        '使用邻域平滑': args.use_smoothing,
        '训练TCN': args.train_tcn,
        '测试集比例': args.test_size,
        '随机种子': args.random_state
    }
    
    # 如果使用SVM分类器，添加SVM特定参数
    if hasattr(args, 'classifier') and args.classifier == 'svm' or not hasattr(args, 'classifier'):
        experiment_data['SVM C参数'] = args.svm_c
        experiment_data['SVM核函数'] = args.svm_kernel
    
    # 如果有TCN训练历史，添加TCN相关数据
    if tcn_history is not None and 'loss' in tcn_history:
        experiment_data['TCN最终Loss'] = round(tcn_history['loss'][-1], 6)
        experiment_data['TCN最佳Loss'] = round(min(tcn_history['loss']), 6)
        # 记录最佳Loss出现的轮次
        best_epoch = tcn_history['loss'].index(min(tcn_history['loss'])) + 1
        experiment_data['TCN最佳Loss轮次'] = best_epoch
    
    # 如果有更多评估指标，添加到数据字典
    if metrics is not None:
        if 'auc_score' in metrics:
            experiment_data['AUC Score(%)'] = round(metrics['auc_score'], 2)
        if 'precision' in metrics:
            experiment_data['Precision(%)'] = round(metrics['precision'], 2)
        if 'recall' in metrics:
            experiment_data['Recall(%)'] = round(metrics['recall'], 2)
        if 'f1_score' in metrics:
            experiment_data['F1 Score(%)'] = round(metrics['f1_score'], 2)
    
    # 检查文件是否存在
    if os.path.exists(excel_file):
        # 读取现有数据
        try:
            existing_df = pd.read_excel(excel_file)
            # 添加新行
            new_df = pd.concat([existing_df, pd.DataFrame([experiment_data])], ignore_index=True)
        except Exception as e:
            print(f"读取Excel文件出错: {e}")
            print("创建新文件...")
            new_df = pd.DataFrame([experiment_data])
    else:
        # 创建新的DataFrame
        new_df = pd.DataFrame([experiment_data])
    
    # 保存到Excel
    try:
        new_df.to_excel(excel_file, index=False, engine='openpyxl')
        format_excel_file(excel_file)  # 格式化Excel文件
        print(f"实验结果已记录到Excel文件: {excel_file}")
    except Exception as e:
        print(f"保存Excel文件出错: {e}")
        # 备用方案：保存为CSV
        csv_file = excel_file.replace('.xlsx', '.csv')
        new_df.to_csv(csv_file, index=False)
        print(f"实验结果已保存为CSV文件: {csv_file}")

def format_excel_file(excel_file):
    """
    格式化Excel文件，使其更易于阅读
    
    参数:
    excel_file - Excel文件路径
    """
    try:
        # 加载工作簿
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # 设置列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # 设置标题行样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置数据单元格对齐方式
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 冻结标题行
        ws.freeze_panes = "A2"
        
        # 保存格式化后的文件
        wb.save(excel_file)
    except Exception as e:
        print(f"格式化Excel文件出错: {e}")